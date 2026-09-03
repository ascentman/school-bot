"""Експорт місячного табеля в XLSX."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from school_bot.reports.matrix import UA_METRIC, MonthMatrix

HEADER_FILL = PatternFill("solid", fgColor="DDE5F0")
WEEKEND_FILL = PatternFill("solid", fgColor="EFEFEF")
OFF_FILL = PatternFill("solid", fgColor="FFF2CC")
MISSING_FILL = PatternFill("solid", fgColor="FBD5D5")
TOTAL_FILL = PatternFill("solid", fgColor="E8F0E4")

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FIRST_DATA_COL = 2


def sheet_title(matrix: MonthMatrix) -> str:
    """Назва аркуша: «2026-09», «2026-09 відсутні», «2026-09 хворі».

    Суфікс, а не окрема схема іменування: базова назва лишається сортованою
    й такою самою, як була, тож аркуш харчування нікуди не «переїжджає».
    """
    base = f"{matrix.year}-{matrix.month:02d}"
    suffix = UA_METRIC[matrix.metric]
    return f"{base} {suffix}" if suffix else base


def _fill_sheet(ws, matrix: MonthMatrix) -> None:

    total_col = FIRST_DATA_COL + len(matrix.columns)
    last_col_letter = get_column_letter(total_col)

    # --- шапка ---
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = matrix.heading
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    if matrix.school_name:
        ws.merge_cells(f"A2:{last_col_letter}2")
        ws["A2"] = matrix.school_name
        ws["A2"].alignment = Alignment(horizontal="center")

    day_row, wd_row, first_data_row = 4, 5, 6

    ws.cell(day_row, 1, "Клас").font = Font(bold=True)
    ws.merge_cells(start_row=day_row, start_column=1, end_row=wd_row, end_column=1)
    ws.cell(day_row, 1).alignment = Alignment(horizontal="center", vertical="center")

    for i, col in enumerate(matrix.columns):
        c = FIRST_DATA_COL + i
        head = ws.cell(day_row, c, col.date.day)
        sub = ws.cell(wd_row, c, col.off_marker or col.weekday_short)
        for cell in (head, sub):
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            if col.is_weekend:
                cell.fill = WEEKEND_FILL
            elif col.off_kind is not None:
                cell.fill = OFF_FILL
            else:
                cell.fill = HEADER_FILL

    tcell = ws.cell(day_row, total_col, "Разом")
    tcell.font = Font(bold=True, size=9)
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    tcell.fill = TOTAL_FILL
    ws.merge_cells(start_row=day_row, start_column=total_col, end_row=wd_row, end_column=total_col)

    # --- рядки класів ---
    for r, row in enumerate(matrix.rows, start=first_data_row):
        name_cell = ws.cell(r, 1, row.name)
        name_cell.font = Font(bold=True)
        name_cell.border = BORDER

        for i, col in enumerate(matrix.columns):
            c = FIRST_DATA_COL + i
            value = row.value(col.date)
            cell = ws.cell(r, c, value if value is not None else None)
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            if col.is_weekend:
                cell.fill = WEEKEND_FILL
            elif col.off_kind is not None:
                cell.fill = OFF_FILL
            elif matrix.is_gap(row, col):
                cell.fill = MISSING_FILL  # пропущений навчальний день

        total_cell = ws.cell(r, total_col, row.total)
        total_cell.font = Font(bold=True)
        total_cell.fill = TOTAL_FILL
        total_cell.alignment = Alignment(horizontal="center")
        total_cell.border = BORDER

    # --- підсумковий рядок ---
    tr = first_data_row + len(matrix.rows)
    ws.cell(tr, 1, "Разом").font = Font(bold=True)
    ws.cell(tr, 1).fill = TOTAL_FILL
    ws.cell(tr, 1).border = BORDER

    for i, col in enumerate(matrix.columns):
        c = FIRST_DATA_COL + i
        # Без огляду на is_school_day: якщо в рядках класів цифра є,
        # підсумок мусить її враховувати, інакше звіт суперечить сам собі.
        cell = ws.cell(tr, c, matrix.day_total(col.date))
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    grand = ws.cell(tr, total_col, matrix.grand_total)
    grand.font = Font(bold=True, size=12)
    grand.fill = TOTAL_FILL
    grand.alignment = Alignment(horizontal="center")
    grand.border = BORDER

    # --- підпис ---
    sign_row = tr + 2
    ws.cell(sign_row, 1, "Відповідальна особа: ______________________")
    ws.cell(sign_row + 1, 1, "Дата: ______________")

    # --- геометрія ---
    ws.freeze_panes = ws.cell(first_data_row, FIRST_DATA_COL)
    ws.column_dimensions["A"].width = 10
    for i in range(len(matrix.columns)):
        ws.column_dimensions[get_column_letter(FIRST_DATA_COL + i)].width = 4.2
    ws.column_dimensions[last_col_letter].width = 8
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_workbook(*matrices: MonthMatrix) -> Workbook:
    """Книга з аркушем на метрику. Перший аркуш — харчування, як і раніше."""
    wb = Workbook()
    for i, matrix in enumerate(matrices):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = sheet_title(matrix)
        _fill_sheet(ws, matrix)
    return wb


def render_xlsx(*matrices: MonthMatrix) -> bytes:
    buf = BytesIO()
    build_workbook(*matrices).save(buf)
    return buf.getvalue()
