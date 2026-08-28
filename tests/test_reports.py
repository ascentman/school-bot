from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from school_bot.db.models import DayKind
from school_bot.domain.calendar import mark_range
from school_bot.domain.meals import upsert_entry
from school_bot.reports.matrix import build_month_matrix
from school_bot.reports.pdf import render_pdf
from school_bot.reports.xlsx import render_xlsx


async def _filled(session, classes, teacher):
    await mark_range(session, date(2026, 9, 14), date(2026, 9, 15), DayKind.VACATION)
    for day, counts in {1: (24, 18, 20), 2: (22, 19, 21), 3: (25, 18, 20)}.items():
        for cls, n in zip(classes, counts, strict=True):
            await upsert_entry(
                session, class_id=cls.id, d=date(2026, 9, day), eating_count=n,
                teacher_id=teacher.id,
            )
    return await build_month_matrix(session, 2026, 9, school_name="Ліцей №1")


async def test_xlsx_structure_and_totals(session, classes, teacher):
    matrix = await _filled(session, classes, teacher)
    ws = load_workbook(BytesIO(render_xlsx(matrix))).active

    assert ws.title == "2026-09"
    assert ws["A1"].value == "Облік харчування учнів — вересень 2026"
    assert ws["A2"].value == "Ліцей №1"

    # шапка: колонка B = 1 вересня, остання колонка даних = 30 вересня
    assert ws.cell(4, 2).value == 1
    assert ws.cell(4, 31).value == 30
    assert ws.cell(4, 32).value == "Разом"

    # 1-А, 1 вересня
    assert ws.cell(6, 2).value == 24
    assert ws.cell(6, 32).value == 71          # 24+22+25
    assert ws.cell(9, 1).value == "Разом"
    assert ws.cell(9, 2).value == 62           # 24+18+20
    assert ws.cell(9, 5).value is None         # 04.09 — даних немає, а не нуль
    assert ws.cell(9, 32).value == matrix.grand_total == 187

    assert ws.freeze_panes == "B6"


async def test_xlsx_marks_weekend_vacation_and_missing(session, classes, teacher):
    matrix = await _filled(session, classes, teacher)
    ws = load_workbook(BytesIO(render_xlsx(matrix))).active

    def fill(row: int, day: int) -> str:
        return ws.cell(row, 1 + day).fill.fgColor.rgb

    assert fill(6, 5) == "00EFEFEF"    # субота 05.09 — сірий
    assert fill(6, 14) == "00FFF2CC"   # канікули 14.09 — жовтий
    assert fill(6, 4) == "00FBD5D5"    # пропущений навчальний день 04.09 — червоний
    assert ws.cell(5, 15).value == "кн"


async def test_xlsx_empty_month_still_renders(session, classes):
    matrix = await build_month_matrix(session, 2026, 9, school_name="Ліцей №1")
    ws = load_workbook(BytesIO(render_xlsx(matrix))).active
    assert ws.cell(9, 32).value == 0


async def test_pdf_renders(session, classes, teacher):
    matrix = await _filled(session, classes, teacher)
    data = render_pdf(matrix)
    assert data.startswith(b"%PDF-")
    assert len(data) > 2000


async def test_pdf_empty_month(session, classes):
    matrix = await build_month_matrix(session, 2026, 2)
    assert render_pdf(matrix).startswith(b"%PDF-")


async def test_xlsx_does_not_flag_future_days(session, classes, teacher):
    await upsert_entry(
        session, class_id=classes[0].id, d=date(2026, 9, 1), eating_count=24, teacher_id=teacher.id
    )
    matrix = await build_month_matrix(session, 2026, 9, today=date(2026, 9, 2))
    ws = load_workbook(BytesIO(render_xlsx(matrix))).active

    def fill(row: int, day: int) -> str:
        return ws.cell(row, 1 + day).fill.fgColor.rgb

    assert fill(6, 2) == "00FBD5D5"    # 02.09 минуло без запису — червоний
    assert fill(6, 3) != "00FBD5D5"    # 03.09 ще не настало — без заливки
    assert fill(6, 30) != "00FBD5D5"
