"""Перевірки побудови Google-таблиці — без мережі й без облікових даних Google.

Модуль будує ту саму сітку, що XLSX і PDF. Обіцянка «три формати не можуть
розійтися» досі трималася лише на тому, що всі троє читають MonthMatrix —
тут вона перевіряється явно.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from school_bot.db.models import DayKind
from school_bot.domain.calendar import mark_range
from school_bot.domain.meals import upsert_entry
from school_bot.reports.matrix import build_month_matrix
from school_bot.reports.sheets import (
    SUMMARY_TAB,
    format_requests,
    matrix_to_grid,
    safe_rebuild_month,
    summary_grid,
    tab_name,
)
from school_bot.reports.xlsx import render_xlsx

HEADER_DAYS_ROW = 3      # 0-based індекс у сітці
HEADER_WD_ROW = 4
FIRST_DATA_ROW = 5


async def _september(session, classes, teacher, *, today=None):
    await mark_range(session, date(2026, 9, 14), date(2026, 9, 15), DayKind.VACATION)
    for day, counts in {1: (24, 18, 20), 2: (22, 19, 21), 3: (25, 18, 20)}.items():
        for cls, n in zip(classes, counts, strict=True):
            await upsert_entry(
                session, class_id=cls.id, d=date(2026, 9, day), eating_count=n,
                teacher_id=teacher.id,
            )
    return await build_month_matrix(session, 2026, 9, school_name="Ліцей №1", today=today)


def test_tab_name_sorts_chronologically():
    """Назва вкладки має сортуватися як рядок — інакше жовтень стане перед лютим."""
    assert tab_name(2026, 9) == "2026-09"
    assert tab_name(2026, 12) == "2026-12"
    assert sorted([tab_name(2026, 10), tab_name(2026, 2)]) == ["2026-02", "2026-10"]


async def test_grid_shape(session, classes, teacher):
    grid = matrix_to_grid(await _september(session, classes, teacher))

    assert grid[0][0] == "Облік харчування учнів — вересень 2026"
    assert grid[1][0] == "Ліцей №1"
    assert grid[HEADER_DAYS_ROW][0] == "Клас"
    assert grid[HEADER_DAYS_ROW][1] == "1"
    assert grid[HEADER_DAYS_ROW][30] == "30"
    assert grid[HEADER_DAYS_ROW][-1] == "Разом"
    assert len(grid) == FIRST_DATA_ROW + len(classes) + 1  # +1 підсумковий рядок


async def test_grid_values_and_totals(session, classes, teacher):
    grid = matrix_to_grid(await _september(session, classes, teacher))

    first_class = grid[FIRST_DATA_ROW]
    assert first_class[0] == "1-А"
    assert first_class[1] == 24          # 1 вересня
    assert first_class[-1] == 71         # 24+22+25

    totals = grid[-1]
    assert totals[0] == "Разом"
    assert totals[1] == 62               # 24+18+20
    assert totals[-1] == 187


async def test_missing_day_is_blank_not_zero(session, classes, teacher):
    """Порожня клітинка і справжній нуль — різні речі у звіті для перевірки."""
    grid = matrix_to_grid(await _september(session, classes, teacher))

    assert grid[FIRST_DATA_ROW][4] == ""   # 4 вересня — даних немає
    assert grid[-1][4] == ""               # і в підсумку теж порожньо


async def test_vacation_and_weekend_markers(session, classes, teacher):
    grid = matrix_to_grid(await _september(session, classes, teacher))
    weekdays = grid[HEADER_WD_ROW]

    assert weekdays[14] == "кн"    # 14 вересня — канікули
    assert weekdays[5] == "Сб"     # 5 вересня — субота
    assert weekdays[1] == "Вт"


async def test_grid_matches_xlsx_cell_for_cell(session, classes, teacher):
    """Головна перевірка: Google-таблиця й XLSX мають збігатися дослівно.

    Якщо хтось змінить один рендерер і забуде інший, перевірка побачить
    у таблиці одні цифри, а у файлі — інші.
    """
    matrix = await _september(session, classes, teacher)
    grid = matrix_to_grid(matrix)
    ws = load_workbook(BytesIO(render_xlsx(matrix))).active

    # XLSX має шапку на рядках 4-5 і дані з 6-го; сітка — те саме, 0-based.
    for col in range(1, len(matrix.columns) + 2):
        assert str(ws.cell(4, col + 1).value or "") == str(grid[HEADER_DAYS_ROW][col])

    for r in range(len(matrix.rows) + 1):          # рядки класів + підсумок
        for col in range(len(matrix.columns) + 2):
            xlsx_value = ws.cell(6 + r, col + 1).value
            grid_value = grid[FIRST_DATA_ROW + r][col]
            assert str(xlsx_value if xlsx_value is not None else "") == str(grid_value), (
                f"розбіжність у рядку {r}, колонці {col}"
            )


async def test_format_requests_cover_the_grid(session, classes, teacher):
    matrix = await _september(session, classes, teacher, today=date(2026, 9, 30))
    reqs = format_requests(matrix, sheet_id=7)

    assert all(
        r.get("repeatCell", {}).get("range", {}).get("sheetId", 7) == 7
        for r in reqs
        if "repeatCell" in r
    )

    frozen = next(r for r in reqs if "updateSheetProperties" in r)
    grid_props = frozen["updateSheetProperties"]["properties"]["gridProperties"]
    assert grid_props["frozenRowCount"] == FIRST_DATA_ROW
    assert grid_props["frozenColumnCount"] == 1


async def test_format_requests_mark_gaps(session, classes, teacher):
    """Пропущені навчальні дні мають бути залиті, майбутні — ні."""
    matrix = await _september(session, classes, teacher, today=date(2026, 9, 3))
    reqs = format_requests(matrix, sheet_id=1)

    red = {"red": 0.984, "green": 0.835, "blue": 0.835}
    painted = {
        (r["repeatCell"]["range"]["startRowIndex"], r["repeatCell"]["range"]["startColumnIndex"])
        for r in reqs
        if r.get("repeatCell", {}).get("cell", {}).get("userEnteredFormat", {}).get(
            "backgroundColor"
        )
        == red
    }
    assert painted == set(), "3 вересня дані є, пізніші дні ще не настали"

    later = format_requests(
        await _september(session, classes, teacher, today=date(2026, 9, 7)), sheet_id=1
    )
    reds = [
        r for r in later
        if r.get("repeatCell", {}).get("cell", {}).get("userEnteredFormat", {}).get(
            "backgroundColor"
        )
        == red
    ]
    assert reds, "4 і 7 вересня минули без записів — мали б підсвітитися"


async def test_summary_grid(session, classes, teacher):
    september = await _september(session, classes, teacher)
    october = await build_month_matrix(session, 2026, 10, school_name="Ліцей №1")

    grid = summary_grid([october, september])

    assert grid[0] == ["Місяць", "1-А", "3-Б", "5-В", "Разом"]
    assert grid[1][0] == "вересень 2026"      # місяці впорядковані
    assert grid[1][-1] == 187
    assert grid[2][0] == "жовтень 2026"
    assert grid[2][-1] == 0


def test_summary_grid_handles_no_data():
    """Без жодного місяця лишається сама шапка, а не порожній список."""
    assert summary_grid([]) == [["Місяць", "Разом"]]


async def test_sync_is_skipped_when_disabled(session, classes, teacher):
    """Без облікових даних Google синк мовчки нічого не робить, а не падає."""
    matrix = await _september(session, classes, teacher)
    assert await safe_rebuild_month(matrix) is None


def test_summary_tab_name_is_stable():
    """Назва вкладки — частина того, що бачить перевірка; змінювати не можна."""
    assert SUMMARY_TAB == "Зведення"


@pytest.mark.parametrize("month,expected", [(1, "2026-01"), (12, "2026-12")])
def test_tab_name_pads_month(month: int, expected: str):
    assert tab_name(2026, month) == expected


async def test_format_requests_paint_weekends_and_vacations(session, classes, teacher):
    """Заливка вихідних і канікул — інакше переплутану умову ніхто не помітить.

    Червону підсвітку пропущених днів перевіряє тест вище; сіра та жовта
    лишалися неперевіреними, хоча саме на них перевірка дивиться першою.
    """
    matrix = await _september(session, classes, teacher, today=date(2026, 9, 30))
    reqs = format_requests(matrix, sheet_id=1)

    grey = {"red": 0.937, "green": 0.937, "blue": 0.937}
    yellow = {"red": 1.0, "green": 0.949, "blue": 0.8}

    def painted_columns(color: dict[str, float]) -> set[int]:
        return {
            r["repeatCell"]["range"]["startColumnIndex"]
            for r in reqs
            if r.get("repeatCell", {})
            .get("cell", {})
            .get("userEnteredFormat", {})
            .get("backgroundColor")
            == color
        }

    # Колонка = день місяця (перша колонка — назви класів).
    weekends = painted_columns(grey)
    assert 5 in weekends and 6 in weekends      # субота й неділя
    assert 1 not in weekends                    # вівторок

    # 14 і 15 вересня 2026 — понеділок і вівторок, тож обидва потрапили
    # в канікули; mark_range вихідні пропускає, вони й так неробочі.
    vacations = painted_columns(yellow)
    assert {14, 15} <= vacations
    assert not (weekends & vacations), "день не може бути водночас сірим і жовтим"


def test_tab_name_suffixes_only_the_extra_metrics():
    """Вкладка харчування має лишитися «2026-09» — її вже бачила перевірка."""
    from school_bot.db.models import MealField
    from school_bot.reports.sheets import tab_name

    assert tab_name(2026, 9) == "2026-09"
    assert tab_name(2026, 9, MealField.EATING) == "2026-09"
    assert tab_name(2026, 9, MealField.ABSENT) == "2026-09 відсутні"
    assert tab_name(2026, 9, MealField.SICK) == "2026-09 хворі"


async def test_absent_grid_carries_the_absent_numbers(session, classes):
    from school_bot.db.models import MealField
    from school_bot.domain.meals import upsert_entry
    from school_bot.reports.matrix import build_month_matrix
    from school_bot.reports.sheets import matrix_to_grid

    await upsert_entry(
        session, class_id=classes[0].id, d=date(2026, 9, 1), eating_count=24,
        absent_count=3, teacher_id=None,
    )
    m = await build_month_matrix(session, 2026, 9, metric=MealField.ABSENT)
    grid = matrix_to_grid(m)

    assert "відсутні" in grid[0][0]
    assert grid[FIRST_DATA_ROW][1] == 3
